#!/usr/bin/env python3
"""Convert Excel ledger to Fidra SQLite database."""

import sqlite3
import uuid
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl


def create_schema(conn: sqlite3.Connection) -> None:
    """Create the Fidra database schema."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id TEXT PRIMARY KEY,
            date TEXT NOT NULL,
            description TEXT NOT NULL,
            amount TEXT NOT NULL,
            type TEXT NOT NULL CHECK (type IN ('income', 'expense')),
            status TEXT NOT NULL CHECK (status IN ('--', 'pending', 'approved', 'rejected', 'planned')),
            sheet TEXT NOT NULL,
            category TEXT,
            party TEXT,
            notes TEXT,
            version INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            modified_at TEXT,
            modified_by TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_transactions_date ON transactions(date);
        CREATE INDEX IF NOT EXISTS idx_transactions_sheet ON transactions(sheet);
        CREATE INDEX IF NOT EXISTS idx_transactions_type ON transactions(type);
        CREATE INDEX IF NOT EXISTS idx_transactions_status ON transactions(status);

        CREATE TABLE IF NOT EXISTS planned_templates (
            id TEXT PRIMARY KEY,
            start_date TEXT NOT NULL,
            description TEXT NOT NULL,
            amount TEXT NOT NULL,
            type TEXT NOT NULL CHECK (type IN ('income', 'expense')),
            frequency TEXT NOT NULL CHECK (frequency IN ('once', 'weekly', 'biweekly', 'monthly', 'quarterly', 'yearly')),
            target_sheet TEXT NOT NULL,
            category TEXT,
            party TEXT,
            end_date TEXT,
            occurrence_count INTEGER,
            skipped_dates TEXT DEFAULT '[]',
            fulfilled_dates TEXT DEFAULT '[]',
            version INTEGER DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_planned_start ON planned_templates(start_date);
        CREATE INDEX IF NOT EXISTS idx_planned_target ON planned_templates(target_sheet);

        CREATE TABLE IF NOT EXISTS sheets (
            id TEXT PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            is_virtual INTEGER DEFAULT 0,
            is_planned INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_sheets_name ON sheets(name);
    """)
    conn.commit()


def convert_excel_to_db(excel_path: Path, db_path: Path) -> None:
    """Convert Excel ledger to Fidra SQLite database."""
    print(f"Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Create database
    print(f"Creating database: {db_path}")
    conn = sqlite3.connect(db_path)
    create_schema(conn)

    now = datetime.now().isoformat()
    sheets_created = set()
    transactions_count = 0

    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        print(f"\nProcessing sheet: {sheet_name}")
        ws = wb[sheet_name]

        # Skip empty sheets or special sheets
        if ws.max_row < 2:
            print(f"  Skipping empty sheet")
            continue

        # Get header row to understand column mapping
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else "")

        print(f"  Headers found: {headers}")

        # Try to find relevant columns
        date_col = None
        desc_col = None
        amount_col = None
        type_col = None
        category_col = None
        party_col = None
        notes_col = None
        status_col = None
        income_col = None
        expense_col = None

        for i, h in enumerate(headers):
            if h in ("date", "transaction date", "trans date"):
                date_col = i
            elif h in ("description", "desc", "details", "transaction", "memo"):
                desc_col = i
            elif h in ("amount", "value", "sum"):
                amount_col = i
            elif h in ("type", "transaction type", "trans type"):
                type_col = i
            elif h in ("category", "cat", "classification"):
                category_col = i
            elif h in ("party", "payee", "payer", "vendor", "customer", "from/to"):
                party_col = i
            elif h in ("notes", "note", "comments", "comment"):
                notes_col = i
            elif h in ("status", "approval", "state"):
                status_col = i
            elif h in ("income", "credit", "in", "credits"):
                income_col = i
            elif h in ("expense", "debit", "out", "debits", "expenses"):
                expense_col = i

        # If no date column found, skip this sheet
        if date_col is None:
            print(f"  No date column found, skipping sheet")
            continue

        # Create sheet record if not exists
        fidra_sheet_name = sheet_name if sheet_name != "Sheet1" else "Main"
        if fidra_sheet_name not in sheets_created:
            conn.execute(
                "INSERT INTO sheets (id, name, is_virtual, is_planned, created_at) VALUES (?, ?, 0, 0, ?)",
                (str(uuid.uuid4()), fidra_sheet_name, now)
            )
            sheets_created.add(fidra_sheet_name)
            print(f"  Created sheet: {fidra_sheet_name}")

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Get date
            date_val = row[date_col] if date_col < len(row) else None
            if date_val is None:
                continue

            # Parse date
            if isinstance(date_val, datetime):
                trans_date = date_val.date().isoformat()
            elif isinstance(date_val, str):
                try:
                    # Try various date formats
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                        try:
                            trans_date = datetime.strptime(date_val, fmt).date().isoformat()
                            break
                        except ValueError:
                            continue
                    else:
                        print(f"  Row {row_idx}: Could not parse date '{date_val}', skipping")
                        continue
                except Exception:
                    print(f"  Row {row_idx}: Could not parse date '{date_val}', skipping")
                    continue
            else:
                print(f"  Row {row_idx}: Unexpected date type {type(date_val)}, skipping")
                continue

            # Get description
            description = ""
            if desc_col is not None and desc_col < len(row) and row[desc_col]:
                description = str(row[desc_col]).strip()
            if not description:
                description = "Unknown transaction"

            # Get amount and type
            amount = Decimal("0")
            trans_type = "expense"

            if income_col is not None and expense_col is not None:
                # Separate income/expense columns
                income_val = row[income_col] if income_col < len(row) else None
                expense_val = row[expense_col] if expense_col < len(row) else None

                if income_val and income_val != 0:
                    try:
                        amount = abs(Decimal(str(income_val).replace(",", "").replace("£", "").replace("$", "").strip()))
                        trans_type = "income"
                    except Exception:
                        pass
                elif expense_val and expense_val != 0:
                    try:
                        amount = abs(Decimal(str(expense_val).replace(",", "").replace("£", "").replace("$", "").strip()))
                        trans_type = "expense"
                    except Exception:
                        pass
            elif amount_col is not None and amount_col < len(row):
                # Single amount column
                amt_val = row[amount_col]
                if amt_val is not None:
                    try:
                        amt_str = str(amt_val).replace(",", "").replace("£", "").replace("$", "").strip()
                        amt_decimal = Decimal(amt_str)
                        amount = abs(amt_decimal)
                        # Determine type from sign or type column
                        if type_col is not None and type_col < len(row) and row[type_col]:
                            type_val = str(row[type_col]).lower().strip()
                            if type_val in ("income", "credit", "in"):
                                trans_type = "income"
                            else:
                                trans_type = "expense"
                        elif amt_decimal > 0:
                            trans_type = "income"
                        else:
                            trans_type = "expense"
                    except Exception:
                        pass

            # Skip zero-amount transactions
            if amount == 0:
                continue

            # Get optional fields
            category = None
            if category_col is not None and category_col < len(row) and row[category_col]:
                category = str(row[category_col]).strip()

            party = None
            if party_col is not None and party_col < len(row) and row[party_col]:
                party = str(row[party_col]).strip()

            notes = None
            if notes_col is not None and notes_col < len(row) and row[notes_col]:
                notes = str(row[notes_col]).strip()

            # Determine status (default to auto-approved for income, pending for expense)
            status = "--" if trans_type == "income" else "pending"
            if status_col is not None and status_col < len(row) and row[status_col]:
                status_val = str(row[status_col]).lower().strip()
                if status_val in ("approved", "yes", "y", "done"):
                    status = "approved"
                elif status_val in ("rejected", "no", "n"):
                    status = "rejected"
                elif status_val in ("pending", "review"):
                    status = "pending"

            # Insert transaction
            trans_id = str(uuid.uuid4())
            conn.execute(
                """INSERT INTO transactions
                   (id, date, description, amount, type, status, sheet, category, party, notes, version, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)""",
                (trans_id, trans_date, description, str(amount), trans_type, status,
                 fidra_sheet_name, category, party, notes, now)
            )
            transactions_count += 1

        print(f"  Processed rows from {sheet_name}")

    # Ensure Main sheet exists
    if "Main" not in sheets_created:
        conn.execute(
            "INSERT INTO sheets (id, name, is_virtual, is_planned, created_at) VALUES (?, ?, 0, 0, ?)",
            (str(uuid.uuid4()), "Main", now)
        )

    conn.commit()
    conn.close()

    print(f"\n✓ Conversion complete!")
    print(f"  Total transactions: {transactions_count}")
    print(f"  Sheets created: {len(sheets_created)}")
    print(f"  Database saved to: {db_path}")


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python convert_excel_to_db.py <excel_file> [output_db]")
        print("Example: python convert_excel_to_db.py ledger.xlsx fidra.db")
        sys.exit(1)

    excel_file = Path(sys.argv[1])
    if not excel_file.exists():
        print(f"Error: Excel file not found: {excel_file}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        db_file = Path(sys.argv[2])
    else:
        db_file = excel_file.with_suffix(".db")

    convert_excel_to_db(excel_file, db_file)
